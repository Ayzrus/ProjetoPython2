import PySimpleGUI as Sg
import openpyxl as xls

Sg.theme('SandyBeach')

exel = xls.Workbook()

std = exel['Sheet']

exel.remove(std)

exel.create_sheet('Teste')

teste_page = exel['Teste']

teste_page["A1"] = "Produtos"

teste_page["B1"] = "Quantidade"

teste_page["C1"] = "Preço"


def main():
    layout = [
        [Sg.Text('Digite 1 para criar folhas\nDigite 2 para registar produtos')],
        [Sg.InputText()],
        [Sg.Submit('Continuar'), Sg.Cancel('Cancelar')]
    ]

    window = Sg.Window("Teste", layout)

    while True:
        event, values = window.read()
        if event == "Cancelar" or event == Sg.WIN_CLOSED:
            break
        elif values[0] == '1' and event == 'Continuar':
            window.hide()
            open_window()
        elif values[0] == '2' and event == 'Continuar':
            window.hide()
            open_window2()
    window.close()


def open_window():
    layout2 = [
        [Sg.Text('Digite o nome da nova folha de exel')],
        [Sg.InputText()],
        [Sg.Submit('Continuar'), Sg.Cancel('Cancelar')]
    ]
    window2 = Sg.Window("Teste", layout2, modal=True)
    choice = None
    while True:
        event1, values1 = window2.read()
        if event1 == "Cancelar" or event1 == Sg.WIN_CLOSED:
            break
        elif event1 == 'Continuar':
            exel.create_sheet(values1[0])
            window2.close()
            main()
    window2.close()


def open_window2():
    layout3 = [
        [Sg.Text('Digite o nome do produto')],
        [Sg.InputText()],
        [Sg.Text('Digite a quantidade do produto')],
        [Sg.InputText()],
        [Sg.Text('Digite o preço do produto')],
        [Sg.InputText()],
        [Sg.Submit('Continuar'), Sg.Cancel('Cancelar')]
    ]
    window3 = Sg.Window("Teste", layout3, modal=True)
    choice = None
    while True:
        event1, values1 = window3.read()
        if event1 == "Cancelar" or event1 == Sg.WIN_CLOSED:
            break
        elif event1 == 'Continuar':
            data = [{
                1: str(values1[0]),
                2:str(values1[1]),
                3:str(values1[2])
            }]
            for i in data:
                print(i)
                teste_page.append(i)
            exel.save('testes.xlsx')
    window3.close()


if __name__ == "__main__":
    main()

